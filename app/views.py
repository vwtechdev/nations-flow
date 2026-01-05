from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.db.models import Sum, Q, Count
from django.db import connection
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.views.decorators.http import require_http_methods

from django.utils import timezone
from datetime import datetime, timedelta, date
from .models import Church, User, Field, Shepherd, Category, Transaction, AccessLog, Notification
from .forms import (
    ChurchForm, UserForm, FieldForm, ShepherdForm,
    CategoryForm, TransactionForm, ChangePasswordForm, EmailAuthenticationForm, NotificationForm
)
from .decorators import admin_required, treasurer_required, admin_or_treasurer_required, password_changed_required
from calendar import monthrange

# Importações para PDF
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Função helper para obter transações baseado no role do usuário
def get_transactions_for_user(user):
    """
    Retorna QuerySet de transações baseado no role do usuário.
    
    - Admin: Todas as transações
    - Tesoureiro: Apenas suas próprias transações
    - Supervisor: Suas próprias transações + transações de tesoureiros que compartilham campos
    """
    if user.is_admin():
        return Transaction.objects.all()
    
    elif user.is_treasurer():
        return Transaction.objects.filter(user=user)
    
    elif user.is_supervisor():
        # Obter campos do supervisor
        supervisor_fields = user.fields.all()
        
        if not supervisor_fields.exists():
            # Se não tem campos, retorna apenas suas próprias transações
            return Transaction.objects.filter(user=user)
        
        # Obter igrejas dos campos do supervisor
        supervisor_churches = Church.objects.filter(field__in=supervisor_fields)
        
        # Obter tesoureiros que compartilham pelo menos um campo com o supervisor
        treasurer_ids = User.objects.filter(
            role='treasurer',
            fields__in=supervisor_fields
        ).distinct().values_list('id', flat=True)
        
        # Retornar: transações próprias + transações de tesoureiros em igrejas dos campos do supervisor
        # Nota: Q já está importado no início do arquivo
        return Transaction.objects.filter(
            Q(user=user) |  # Suas próprias transações
            Q(
                user_id__in=treasurer_ids,
                church__in=supervisor_churches
            )  # Transações de tesoureiros dos mesmos campos
        ).distinct()
    
    else:
        return Transaction.objects.none()

# Views de Autenticação
def login_view(request):
    if request.user.is_authenticated:
        # Redirecionar tesoureiros e supervisores para a lista de transações
        if not request.user.is_admin():
            return redirect('transaction_list')
        # Admin vai para o dashboard
        return redirect('index')
    
    if request.method == 'POST':
        form = EmailAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('username')  # O campo username agora é email
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=email, password=password)
            if user is not None:
                login(request, user)
                
                # Registrar o login no AccessLog
                AccessLog.objects.create(
                    user=user,
                    action='login',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                messages.success(request, f'Bem-vindo, {user.get_full_name()}!')
                # Verifica se o usuário precisa trocar a senha
                if not user.password_changed:
                    return render(request, 'registration/redirect_to_change_password.html', {
                        'title': 'Redirecionando...',
                        'user': user
                    })
                # Se há um parâmetro next, redireciona para ele
                next_url = request.GET.get('next')
                if next_url and next_url != '/change-password/':
                    return redirect(next_url)
                # Redirecionar tesoureiros e supervisores para a lista de transações
                if not user.is_admin():
                    return redirect('transaction_list')
                # Admin vai para o dashboard
                return redirect('index')
    else:
        form = EmailAuthenticationForm()
    
    return render(request, 'registration/login.html', {'form': form, 'title': 'Login'})

def logout_view(request):
    """View para fazer logout do usuário"""
    # Registrar o logout no AccessLog antes de fazer logout
    if request.user.is_authenticated:
        AccessLog.objects.create(
            user=request.user,
            action='logout',
            ip_address=request.META.get('REMOTE_ADDR')
        )
    
    # Fazer logout do usuário
    logout(request)
    
    # Adicionar mensagem de sucesso
    messages.success(request, 'Você saiu do sistema com sucesso!')
    
    # Redirecionar para a página de login
    return redirect('login')

@login_required
def change_password(request):
    """Força o usuário a trocar a senha no primeiro login"""
    if request.user.password_changed:
        # Redirecionar tesoureiros e supervisores para a lista de transações
        if not request.user.is_admin():
            return redirect('transaction_list')
        return redirect('index')
    
    if request.method == 'POST':
        form = ChangePasswordForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            # Fazer logout do usuário após trocar a senha
            logout(request)
            messages.success(request, 'Senha alterada com sucesso! Faça login novamente com sua nova senha.')
            return render(request, 'registration/redirect_to_login.html', {
                'title': 'Redirecionando...'
            })
        else:
            print(f"DEBUG: Form errors: {form.errors}")
    else:
        form = ChangePasswordForm(request.user)
    
    context = {
        'title': 'Alterar Senha',
        'form': form,
        'force_change': True,
    }
    
    return render(request, 'registration/change_password.html', context)

# Views Principais
@password_changed_required
def index(request):
    """Dashboard principal - apenas para administradores"""
    # Redirecionar tesoureiros e supervisores para a lista de transações
    if not request.user.is_admin():
        return redirect('transaction_list')
    
    # Obter filtros da URL
    selected_start_date = request.GET.get('start_date', '')
    selected_end_date = request.GET.get('end_date', '')
    selected_category = request.GET.get('category', '')
    selected_type = request.GET.get('type', '')
    selected_field = request.GET.get('field', '')
    selected_church = request.GET.get('church', '')
    selected_shepherd = request.GET.get('shepherd', '')
    selected_user = request.GET.get('user', '')
    
    # Definir datas padrão se não fornecidas (ano atual)
    today = date.today()
    if not selected_start_date:
        # Primeiro dia do ano atual
        selected_start_date = date(today.year, 1, 1).strftime('%Y-%m-%d')
    if not selected_end_date:
        # Último dia do ano atual
        selected_end_date = date(today.year, 12, 31).strftime('%Y-%m-%d')
    
    # Converter para objetos date para cálculos
    start_date = datetime.strptime(selected_start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(selected_end_date, '%Y-%m-%d').date()
    
    context = {
        'title': 'Dashboard',
        'user': request.user,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
        'selected_category': selected_category,
        'selected_type': selected_type,
        'selected_field': selected_field,
        'selected_church': selected_church,
        'selected_shepherd': selected_shepherd,
        'selected_user': selected_user,
        'categories': Category.objects.all(),
        'current_year': today.year
    }
    
    # Adicionar campos, igrejas, pastores e usuários ao contexto baseado no tipo de usuário
    if request.user.is_admin():
        context['fields'] = Field.objects.all()
        context['churches'] = Church.objects.all()
        context['shepherds'] = Shepherd.objects.all()
        context['users'] = User.objects.exclude(email='vwtechdev@gmail.com').order_by('first_name', 'last_name')
    else:
        # Tesoureiro vê seus campos e suas igrejas
        if request.user.fields.exists():
            context['fields'] = request.user.fields.all()
            user_churches = Church.objects.filter(field__in=request.user.fields.all())
            context['churches'] = user_churches
            context['shepherds'] = Shepherd.objects.filter(church__field__in=request.user.fields.all()).distinct()
        else:
            context['fields'] = Field.objects.none()
            context['churches'] = Church.objects.none()
            context['shepherds'] = Shepherd.objects.none()
        context['users'] = User.objects.none()
    
    # Base de transações
    if request.user.is_admin():
        base_transactions = Transaction.objects.all()
    elif request.user.is_supervisor():
        base_transactions = get_transactions_for_user(request.user)
    else:
        base_transactions = Transaction.objects.filter(user=request.user)
    
    # Aplicar filtros
    filtered_transactions = base_transactions
    
    # Filtro por período de datas
    filtered_transactions = filtered_transactions.filter(date__gte=start_date, date__lte=end_date)
    
    # Filtro por categoria
    if selected_category:
        filtered_transactions = filtered_transactions.filter(category_id=selected_category)
    
    # Filtro por tipo
    if selected_type:
        filtered_transactions = filtered_transactions.filter(type=selected_type)
    
    # Filtro por campo
    if selected_field:
        filtered_transactions = filtered_transactions.filter(church__field_id=selected_field)
    
    # Filtro por igreja
    if selected_church:
        filtered_transactions = filtered_transactions.filter(church_id=selected_church)
    
    # Filtro por pastor
    if selected_shepherd:
        filtered_transactions = filtered_transactions.filter(church__shepherd_id=selected_shepherd)
    
    # Filtro por usuário (apenas para administradores)
    if selected_user and request.user.is_admin():
        filtered_transactions = filtered_transactions.filter(user_id=selected_user)
    
    # Calcular totais
    total_transactions = filtered_transactions.count()
    total_income = filtered_transactions.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
    total_expense = filtered_transactions.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
    balance = total_income - total_expense
    
    # Dados para o gráfico por categoria
    categories_data = []
    categories = Category.objects.all()
    
    for category in categories:
        category_transactions = filtered_transactions.filter(category=category)
        income = category_transactions.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
        expense = category_transactions.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
        
        if income > 0 or expense > 0:
            categories_data.append({
                'category': category.name,
                'income': float(income),
                'expense': float(expense)
            })
    
    # Dados por campo - agrupar igrejas por campo
    churches_data = []
    if request.user.is_admin():
        # Para administradores, mostrar todos os campos
        all_fields = Field.objects.all()
    else:
        # Para tesoureiros, mostrar apenas seus campos
        if request.user.fields.exists():
            all_fields = request.user.fields.all()
        else:
            all_fields = Field.objects.none()
    
    for field in all_fields:
        # Buscar todas as transações das igrejas deste campo
        field_transactions = filtered_transactions.filter(church__field=field)
        
        income = field_transactions.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
        expense = field_transactions.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
        
        # Mostrar apenas campos com valores maiores que zero
        if income > 0 or expense > 0:
            churches_data.append({
                'name': field.name,
                'income': float(income),
                'expense': float(expense)
            })
    
    # Ordenar por nome do campo
    churches_data.sort(key=lambda x: x['name'])
    
    # Dados para o gráfico de Entrada e Saída por Igreja individual
    churches_individual_data = []
    if request.user.is_admin():
        # Para administradores, mostrar todos os campos
        all_churches = Church.objects.all()
    else:
        # Para tesoureiros, mostrar apenas igrejas dos seus campos
        if request.user.fields.exists():
            all_churches = Church.objects.filter(field__in=request.user.fields.all())
        else:
            all_churches = Church.objects.none()
    
    for church in all_churches:
        church_transactions = filtered_transactions.filter(church=church)
        income = church_transactions.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
        expense = church_transactions.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
        
        # Mostrar apenas igrejas com valores maiores que zero
        if income > 0 or expense > 0:
            churches_individual_data.append({
                'name': church.name,
                'field': church.field.name if church.field else 'Sem Campo',
                'income': float(income),
                'expense': float(expense)
            })
    
    # Ordenar por nome da igreja
    churches_individual_data.sort(key=lambda x: x['name'])
    
    # Dados mensais para o gráfico: apenas meses entre start_date e end_date
    monthly_data = []

    # Nomes dos meses em português
    month_names = [
        'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
    ]

    # Construir lista de meses do intervalo
    months_in_period = []
    cursor = date(start_date.year, start_date.month, 1)
    while cursor <= end_date:
        month_start = date(cursor.year, cursor.month, 1)
        # Último dia do mês
        if cursor.month == 12:
            month_end = date(cursor.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(cursor.year, cursor.month + 1, 1) - timedelta(days=1)
        # Cortar para o intervalo selecionado
        effective_start = max(month_start, start_date)
        effective_end = min(month_end, end_date)
        months_in_period.append({
            'year': cursor.year,
            'month': cursor.month,
            'month_start': effective_start,
            'month_end': effective_end
        })
        # Avançar para próximo mês
        if cursor.month == 12:
            cursor = date(cursor.year + 1, 1, 1)
        else:
            cursor = date(cursor.year, cursor.month + 1, 1)

    # Calcular valores por mês
    for month_info in months_in_period:
        month_start = month_info['month_start']
        month_end = month_info['month_end']

        # Base de transações por perfil
        if request.user.is_admin():
            month_transactions = Transaction.objects.all()
        else:
            if request.user.fields.exists():
                user_churches = Church.objects.filter(field__in=request.user.fields.all())
                month_transactions = Transaction.objects.filter(church__in=user_churches)
            else:
                month_transactions = Transaction.objects.none()
        # Filtros do mês e demais filtros
        month_transactions = month_transactions.filter(date__gte=month_start, date__lte=month_end)
        if selected_category:
            month_transactions = month_transactions.filter(category_id=selected_category)
        if selected_type:
            month_transactions = month_transactions.filter(type=selected_type)
        if selected_field:
            month_transactions = month_transactions.filter(church__field_id=selected_field)
        if selected_church:
            month_transactions = month_transactions.filter(church_id=selected_church)
        if selected_shepherd:
            month_transactions = month_transactions.filter(church__shepherd_id=selected_shepherd)
        if selected_user and request.user.is_admin():
            month_transactions = month_transactions.filter(user_id=selected_user)
        month_income = month_transactions.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
        month_expense = month_transactions.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0

        month_balance = float(month_income) - float(month_expense)
        monthly_data.append({
            'month': f"{month_names[month_info['month'] - 1]}/{month_info['year']}",
            'month_number': month_info['month'],
            'year': month_info['year'],
            'income': float(month_income),
            'expense': float(month_expense),
            'balance': month_balance
        })
    
    # Transações recentes (últimas 10)
    recent_transactions = filtered_transactions.order_by('-date')[:10]
    
    # Logs de acesso recentes (apenas para administradores)
    if request.user.is_admin():
        access_logs = AccessLog.objects.select_related('user').exclude(
            user__email='vwtechdev@gmail.com'
        ).order_by('-timestamp')[:20]
    else:
        access_logs = AccessLog.objects.none()
    

    
    # Adicionar dados ao contexto
    context.update({
        'total_transactions': total_transactions,
        'total_income': total_income,
        'total_expense': total_expense,
        'balance': balance,
        'categories_data': categories_data,
        'churches_data': churches_data,
        'churches_individual_data': churches_individual_data,
        'monthly_data': monthly_data,  # Dados mensais para o gráfico de linhas
        'monthly_data_json': json.dumps(monthly_data),
        'recent_transactions': recent_transactions,
        'access_logs': access_logs,
        'selected_category_name': Category.objects.get(id=selected_category).name if selected_category else None,
        'selected_field_name': Field.objects.get(id=selected_field).name if selected_field else None,
        'selected_church_name': Church.objects.get(id=selected_church).name if selected_church else None,
    })
    
    return render(request, 'pages/dashboard.html', context)

# Views de Transações
@password_changed_required
@admin_or_treasurer_required
def transaction_list(request):
    """Lista de transações"""
    
    if request.user.is_admin():
        transactions = Transaction.objects.all()
    elif request.user.is_supervisor():
        transactions = get_transactions_for_user(request.user)
    else:
        # Tesoureiro: ver apenas transações criadas por ele
        transactions = Transaction.objects.filter(user=request.user)
    
    # Filtros
    search = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    transaction_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    selected_field = request.GET.get('field', '')
    selected_church = request.GET.get('church', '')
    selected_shepherd = request.GET.get('shepherd', '')
    selected_user = request.GET.get('user', '')
    
    # Definir datas padrão se não fornecidas
    today = date.today()
    if not date_from:
        # Primeiro dia do mês atual
        first_day = date(today.year, today.month, 1)
        date_from = first_day.strftime('%Y-%m-%d')
    
    if not date_to:
        # Último dia do mês atual
        last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])
        date_to = last_day.strftime('%Y-%m-%d')
    
    if search:
        transactions = transactions.filter(
            Q(desc__icontains=search) | 
            Q(category__name__icontains=search)
        )
    
    if category_id:
        transactions = transactions.filter(category_id=category_id)
    
    if transaction_type:
        transactions = transactions.filter(type=transaction_type)
    
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
    
    # Filtro por campo (para administradores ou tesoureiros com múltiplos campos)
    if selected_field:
        if request.user.is_admin():
            # Administradores podem filtrar por qualquer campo
            transactions = transactions.filter(church__field_id=selected_field)
        elif request.user.fields.count() > 1:
            # Tesoureiros com múltiplos campos podem filtrar por seus campos
            if request.user.fields.filter(id=selected_field).exists():
                transactions = transactions.filter(church__field_id=selected_field)
    
    # Filtro por igreja
    if selected_church:
        transactions = transactions.filter(church_id=selected_church)
    
    # Filtro por pastor
    if selected_shepherd:
        transactions = transactions.filter(church__shepherd_id=selected_shepherd)
    
    # Filtro por usuário (para administradores e supervisores)
    if selected_user:
        if request.user.is_admin():
            transactions = transactions.filter(user_id=selected_user)
        elif request.user.is_supervisor():
            # Supervisor só pode filtrar por usuários que ele pode ver (ele mesmo ou tesoureiros dos mesmos campos)
            supervisor_fields = request.user.fields.all()
            if supervisor_fields.exists():
                allowed_user_ids = User.objects.filter(
                    Q(id=request.user.id) |  # O próprio supervisor
                    Q(role='treasurer', fields__in=supervisor_fields)  # Tesoureiros dos mesmos campos
                ).distinct().values_list('id', flat=True)
                if int(selected_user) in allowed_user_ids:
                    transactions = transactions.filter(user_id=selected_user)
    
    # Calcular totais
    total_transactions = transactions.count()
    total_income = transactions.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
    total_expense = transactions.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
    balance = total_income - total_expense  
    
    # Preparar dados para os filtros
    if request.user.is_admin():
        fields = Field.objects.all()
        churches = Church.objects.all()
        shepherds = Shepherd.objects.all()
        users = User.objects.exclude(email='vwtechdev@gmail.com').order_by('first_name', 'last_name')
        print(f"DEBUG VIEW - Usuários carregados: {users.count()}")
        print(f"DEBUG VIEW - Usuário atual é admin: {request.user.is_admin()}")
        if selected_field:
            churches = churches.filter(field_id=selected_field)
        if selected_shepherd:
            churches = churches.filter(shepherd_id=selected_shepherd)
    else:
        # Verificar se o usuário tem campos associados
        if request.user.fields.exists():
            fields = request.user.fields.all()
            churches = Church.objects.filter(field__in=request.user.fields.all())
            shepherds = Shepherd.objects.filter(church__field__in=request.user.fields.all()).distinct()
            
            # Para Supervisor, incluir usuários que ele pode ver (ele mesmo e tesoureiros dos mesmos campos)
            if request.user.is_supervisor():
                supervisor_fields = request.user.fields.all()
                users = User.objects.filter(
                    Q(id=request.user.id) |  # O próprio supervisor
                    Q(role='treasurer', fields__in=supervisor_fields)  # Tesoureiros dos mesmos campos
                ).distinct().exclude(email='vwtechdev@gmail.com').order_by('first_name', 'last_name')
            else:
                users = User.objects.none()
            
            # Se o usuário tem múltiplos campos, permitir filtro por campo
            if request.user.fields.count() > 1 and selected_field:
                churches = churches.filter(field_id=selected_field)
            if selected_shepherd:
                churches = churches.filter(shepherd_id=selected_shepherd)
        else:
            # Se o usuário não tem campos, mostrar mensagem de erro
            messages.error(request, 'Você não tem campos associados. Entre em contato com o administrador.')
            return redirect('index')
    
    # Nomes dos filtros selecionados
    selected_field_name = fields.get(id=selected_field).name if selected_field and fields.filter(id=selected_field).exists() else None
    selected_church_name = churches.get(id=selected_church).name if selected_church and churches.filter(id=selected_church).exists() else None
    
    context = {
        'title': 'Transações',
        'categories': Category.objects.all(),
        'fields': fields,
        'churches': churches,
        'shepherds': shepherds,
        'users': users,
        'category_id': category_id,
        'selected_category': category_id,
        'transaction_type': transaction_type,
        'selected_type': transaction_type,
        'date_from': date_from,
        'date_to': date_to,
        'selected_field': selected_field,
        'selected_church': selected_church,
        'selected_shepherd': selected_shepherd,
        'selected_user': selected_user,
        'selected_field_name': selected_field_name,
        'selected_church_name': selected_church_name,
        'total_transactions': total_transactions,
        'total_income': total_income,
        'total_expense': total_expense,
        'balance': balance,
        'user': request.user,
    }
    
    return render(request, 'pages/transaction_list.html', context)

# API para paginação AJAX das transações
@password_changed_required
@admin_or_treasurer_required
def transaction_list_api(request):
    """API para listar transações com paginação AJAX"""
    
    if request.user.is_admin():
        transactions = Transaction.objects.all()
        print(f"DEBUG API - Usuário admin: {transactions.count()} transações totais")
    elif request.user.is_supervisor():
        transactions = get_transactions_for_user(request.user)
        print(f"DEBUG API - Usuário supervisor: {transactions.count()} transações de tesoureiros")
    else:
        # Tesoureiro: ver apenas transações criadas por ele
        transactions = Transaction.objects.filter(user=request.user)
        print(f"DEBUG API - Usuário tesoureiro: {transactions.count()} transações do próprio usuário")
    
    # Filtros
    category_id = request.GET.get('category', '')
    transaction_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    selected_field = request.GET.get('field', '')
    selected_church = request.GET.get('church', '')
    selected_shepherd = request.GET.get('shepherd', '')
    selected_user = request.GET.get('user', '')
    
    # Debug para filtros
    print(f"DEBUG API - shepherd: {selected_shepherd}")
    print(f"DEBUG API - user: {selected_user}")
    print(f"DEBUG API - request.GET: {dict(request.GET)}")
    
    # Definir datas padrão se não fornecidas
    today = date.today()
    if not date_from:
        first_day = date(today.year, today.month, 1)
        date_from = first_day.strftime('%Y-%m-%d')
    
    if not date_to:
        last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])
        date_to = last_day.strftime('%Y-%m-%d')
    
    if category_id:
        transactions = transactions.filter(category_id=category_id)
    
    if transaction_type:
        transactions = transactions.filter(type=transaction_type)
    
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
    
    if selected_field:
        if request.user.is_admin():
            transactions = transactions.filter(church__field_id=selected_field)
        elif request.user.fields.count() > 1:
            if request.user.fields.filter(id=selected_field).exists():
                transactions = transactions.filter(church__field_id=selected_field)
    
    if selected_church:
        transactions = transactions.filter(church_id=selected_church)
    
    if selected_shepherd:
        print(f"DEBUG API - Aplicando filtro de pastor: {selected_shepherd}")
        print(f"DEBUG API - Transações antes do filtro: {transactions.count()}")
        
        # Verificar se existem transações com esse pastor
        shepherd_transactions = transactions.filter(church__shepherd_id=selected_shepherd)
        print(f"DEBUG API - Transações com pastor {selected_shepherd}: {shepherd_transactions.count()}")
        
        # Verificar se o pastor existe
        try:
            shepherd = Shepherd.objects.get(id=selected_shepherd)
            print(f"DEBUG API - Pastor encontrado: {shepherd.name}")
        except Shepherd.DoesNotExist:
            print(f"DEBUG API - Pastor com ID {selected_shepherd} não existe!")
        
        transactions = transactions.filter(church__shepherd_id=selected_shepherd)
        print(f"DEBUG API - Após filtro de pastor: {transactions.count()} transações")
    
    if selected_user and request.user.is_admin():
        transactions = transactions.filter(user_id=selected_user)
    
    # Calcular totais
    total_transactions = transactions.count()
    total_income = transactions.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
    total_expense = transactions.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
    balance = total_income - total_expense
    
    # Paginação
    page = int(request.GET.get('page', 1))
    per_page = 50
    start = (page - 1) * per_page
    end = start + per_page
    
    # Obter transações da página atual (evitar N+1)
    page_transactions = transactions.select_related('category', 'church', 'church__field', 'church__shepherd', 'user').order_by('-date')[start:end]
    
    # Preparar dados das transações
    transactions_data = []
    for transaction in page_transactions:
        transactions_data.append({
            'id': transaction.id,
            'date': transaction.date.strftime('%d/%m/%Y'),
            'type': transaction.type,
            'type_display': transaction.get_type_display(),
            'category_name': transaction.category.name,
            'field_name': transaction.church.field.name,
            'church_name': transaction.church.name,
            'shepherd_name': transaction.church.shepherd.name if transaction.church and transaction.church.shepherd else None,
            'desc': transaction.desc or '-',
            'value': float(transaction.value),
            'proof': transaction.proof.url if transaction.proof else None,
            'user_name': (transaction.user.get_full_name() or transaction.user.username) if transaction.user else None,
            'can_edit': request.user.is_admin(),
            'can_view': True,
        })
    
    # Calcular informações de paginação
    total_pages = (total_transactions + per_page - 1) // per_page
    
    return JsonResponse({
        'transactions': transactions_data,
        'pagination': {
            'current_page': page,
            'total_pages': total_pages,
            'per_page': per_page,
            'total_items': total_transactions,
            'has_previous': page > 1,
            'has_next': page < total_pages,
            'previous_page': page - 1 if page > 1 else None,
            'next_page': page + 1 if page < total_pages else None,
        },
        'totals': {
            'total_transactions': total_transactions,
            'total_income': float(total_income),
            'total_expense': float(total_expense),
            'balance': float(balance),
        }
    })

@password_changed_required
@admin_or_treasurer_required
def transaction_summary_api(request):
    """API de resumo para dashboard: retorna agregados e séries sem expor filtros na URL."""
    if request.user.is_admin():
        base_transactions = Transaction.objects.all()
    elif request.user.is_supervisor():
        base_transactions = get_transactions_for_user(request.user)
        if not base_transactions.exists() and not request.user.fields.exists():
            return JsonResponse({'error': 'Supervisor sem campos associados'}, status=400)
    else:
        if request.user.fields.exists():
            user_churches = Church.objects.filter(field__in=request.user.fields.all())
            base_transactions = Transaction.objects.filter(church__in=user_churches)
        else:
            return JsonResponse({'error': 'Usuário sem campos associados'}, status=400)

    # Ler filtros do corpo JSON (POST) ou do GET como fallback, mas sem refletir na URL do front
    if request.method == 'POST':
        try:
            payload = json.loads(request.body or '{}')
        except Exception:
            payload = {}
    else:
        payload = {}

    selected_category = str(payload.get('category', '')).strip()
    selected_type = str(payload.get('type', '')).strip()
    selected_field = str(payload.get('field', '')).strip()
    selected_church = str(payload.get('church', '')).strip()
    selected_shepherd = str(payload.get('shepherd', '')).strip()
    selected_user = str(payload.get('user', '')).strip()
    date_from = str(payload.get('date_from', '')).strip()
    date_to = str(payload.get('date_to', '')).strip()
    monthly_use_current_year = bool(payload.get('monthly_use_current_year', False))

    # Datas padrão: mês atual
    today = date.today()
    if not date_from:
        date_from = date(today.year, today.month, 1).strftime('%Y-%m-%d')
    if not date_to:
        last_day = monthrange(today.year, today.month)[1]
        date_to = date(today.year, today.month, last_day).strftime('%Y-%m-%d')

    filtered = base_transactions
    if date_from:
        filtered = filtered.filter(date__gte=date_from)
    if date_to:
        filtered = filtered.filter(date__lte=date_to)
    if selected_category:
        filtered = filtered.filter(category_id=selected_category)
    if selected_type:
        filtered = filtered.filter(type=selected_type)
    if selected_field:
        if request.user.is_admin():
            filtered = filtered.filter(church__field_id=selected_field)
        elif request.user.fields.count() > 1:
            if request.user.fields.filter(id=selected_field).exists():
                filtered = filtered.filter(church__field_id=selected_field)
    if selected_church:
        filtered = filtered.filter(church_id=selected_church)
    if selected_shepherd:
        filtered = filtered.filter(church__shepherd_id=selected_shepherd)
    if selected_user and request.user.is_admin():
        filtered = filtered.filter(user_id=selected_user)

    # Totais
    total_transactions = filtered.count()
    total_income = filtered.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
    total_expense = filtered.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
    balance = float(total_income) - float(total_expense)

    # Por categoria
    categories_qs = Category.objects.all()
    categories_data = []
    for category in categories_qs:
        cat_qs = filtered.filter(category=category)
        income = cat_qs.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
        expense = cat_qs.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
        if income > 0 or expense > 0:
            categories_data.append({
                'category': category.name,
                'income': float(income),
                'expense': float(expense)
            })

    # Por campo
    if request.user.is_admin():
        fields_qs = Field.objects.all()
    else:
        fields_qs = request.user.fields.all() if request.user.fields.exists() else Field.objects.none()
    churches_data = []
    for field in fields_qs:
        field_qs = filtered.filter(church__field=field)
        income = field_qs.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
        expense = field_qs.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
        if income > 0 or expense > 0:
            churches_data.append({
                'name': field.name,
                'income': float(income),
                'expense': float(expense)
            })
    churches_data.sort(key=lambda x: x['name'])

    # Por igreja
    if request.user.is_admin():
        churches_qs = Church.objects.all()
    else:
        churches_qs = Church.objects.filter(field__in=request.user.fields.all()) if request.user.fields.exists() else Church.objects.none()
    churches_individual_data = []
    for church in churches_qs:
        ch_qs = filtered.filter(church=church)
        income = ch_qs.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
        expense = ch_qs.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
        if income > 0 or expense > 0:
            churches_individual_data.append({
                'name': church.name,
                'field': church.field.name if church.field else 'Sem Campo',
                'income': float(income),
                'expense': float(expense)
            })
    churches_individual_data.sort(key=lambda x: x['name'])

    # Série mensal entre intervalos
    def _month_iter(start_date_str, end_date_str):
        start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        cursor_date = date(start_date_obj.year, start_date_obj.month, 1)
        months = []
        while cursor_date <= end_date_obj:
            if cursor_date.month == 12:
                month_end = date(cursor_date.year, 12, 31)
                next_first = date(cursor_date.year + 1, 1, 1)
            else:
                next_first = date(cursor_date.year, cursor_date.month + 1, 1)
                month_end = next_first - timedelta(days=1)
            eff_start = max(cursor_date, start_date_obj)
            eff_end = min(month_end, end_date_obj)
            months.append((cursor_date.year, cursor_date.month, eff_start, eff_end))
            cursor_date = next_first
        return months

    month_names = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    monthly_data = []
    # Determinar intervalo da série mensal
    if monthly_use_current_year:
        current_year = today.year
        series_start = date(current_year, 1, 1).strftime('%Y-%m-%d')
        series_end = date(current_year, 12, 31).strftime('%Y-%m-%d')
    else:
        series_start = date_from
        series_end = date_to
    for year, month, m_start, m_end in _month_iter(series_start, series_end):
        month_qs = base_transactions
        month_qs = month_qs.filter(date__gte=m_start, date__lte=m_end)
        if selected_category:
            month_qs = month_qs.filter(category_id=selected_category)
        if selected_type:
            month_qs = month_qs.filter(type=selected_type)
        if selected_field:
            if request.user.is_admin():
                month_qs = month_qs.filter(church__field_id=selected_field)
            elif request.user.fields.count() > 1 and request.user.fields.filter(id=selected_field).exists():
                month_qs = month_qs.filter(church__field_id=selected_field)
        if selected_church:
            month_qs = month_qs.filter(church_id=selected_church)
        if selected_shepherd:
            month_qs = month_qs.filter(church__shepherd_id=selected_shepherd)
        if selected_user and request.user.is_admin():
            month_qs = month_qs.filter(user_id=selected_user)
        month_income = month_qs.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
        month_expense = month_qs.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
        monthly_data.append({
            'month': f"{month_names[month-1]}/{year}",
            'month_number': month,
            'year': year,
            'income': float(month_income),
            'expense': float(month_expense),
            'balance': float(month_income) - float(month_expense)
        })

    return JsonResponse({
        'totals': {
            'total_transactions': total_transactions,
            'total_income': float(total_income),
            'total_expense': float(total_expense),
            'balance': balance,
        },
        'categories_data': categories_data,
        'churches_data': churches_data,
        'churches_individual_data': churches_individual_data,
        'monthly_data': monthly_data,
        'filters_applied': {
            'date_from': date_from,
            'date_to': date_to,
            'category': selected_category or None,
            'type': selected_type or None,
            'field': selected_field or None,
            'church': selected_church or None,
            'shepherd': selected_shepherd or None,
            'user': selected_user or None,
        }
    })

@password_changed_required
@admin_or_treasurer_required
def transaction_create(request):
    """Criar nova transação"""
    if request.method == 'POST':
        form = TransactionForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.user = request.user
            transaction.save()
            
            # Verificar se deve criar um lembrete (apenas para administradores)
            if request.user.is_admin() and form.cleaned_data.get('create_reminder'):
                try:
                    # Criar notificação de lembrete (1 dia após a data da transação, às 9h)
                    reminder_date = datetime.combine(transaction.date, datetime.min.time().replace(hour=9, minute=0))
                    reminder_date += timedelta(days=1)
                    reminder_date = timezone.make_aware(reminder_date)
                    
                    Notification.objects.create(
                        title=f"Lembrete: {transaction.category.name}",
                        body=f"Lembrete sobre a transação de {transaction.get_type_display().lower()} de {transaction.get_formatted_value()} para {transaction.church.name}.",
                        date=reminder_date,
                        is_read=False,
                        repeat=False,
                        repeat_frequency='none',
                        created_by=request.user
                    )
                    messages.success(request, 'Transação criada com sucesso e lembrete criado!')
                except Exception as e:
                    messages.success(request, 'Transação criada com sucesso, mas houve um erro ao criar o lembrete.')
                    print(f"Erro ao criar lembrete: {e}")
            else:
                messages.success(request, 'Transação criada com sucesso!')
            
            return redirect('transaction_list')
    else:
        form = TransactionForm(user=request.user)
    
    # Preparar dados para os filtros baseado no tipo de usuário
    if request.user.is_admin():
        fields = Field.objects.all().order_by('name')
        churches = Church.objects.all().order_by('name')
    else:
        # Tesoureiros veem apenas seus campos e suas igrejas
        if request.user.fields.exists():
            fields = request.user.fields.all().order_by('name')
            churches = Church.objects.filter(field__in=request.user.fields.all()).order_by('name')
        else:
            fields = Field.objects.none()
            churches = Church.objects.none()
    
    context = {
        'title': 'Nova Transação',
        'form': form,
        'categories': Category.objects.all().order_by('name'),
        'fields': fields,
        'churches': churches,
    }
    
    return render(request, 'pages/transaction_form.html', context)

@password_changed_required
@admin_or_treasurer_required
def transaction_view(request, pk):
    """Visualizar transação - Modo somente leitura"""
    transaction = get_object_or_404(Transaction, pk=pk)
    
    # Verificar se o usuário tem acesso à transação
    if not request.user.is_admin():
        # Supervisor e Tesoureiro: verificar se têm acesso
        if request.user.is_supervisor():
            # Supervisor pode ver suas próprias transações ou de tesoureiros dos mesmos campos
            supervisor_fields = request.user.fields.all()
            can_view = (
                transaction.user == request.user or  # Sua própria transação
                (
                    transaction.user.role == 'treasurer' and
                    transaction.user.fields.filter(id__in=supervisor_fields.values_list('id', flat=True)).exists() and
                    supervisor_fields.filter(id=transaction.church.field.id).exists()
                )
            )
        else:
            # Tesoureiro: apenas suas próprias transações
            can_view = transaction.user == request.user
        
        if not can_view:
            messages.error(request, 'Você não tem permissão para visualizar esta transação.')
            return redirect('transaction_list')
    
    form = TransactionForm(instance=transaction, user=request.user)
    
    # Desabilitar todos os campos do formulário
    for field_name in form.fields:
        form.fields[field_name].disabled = True
    
    # Preparar dados para os filtros baseado no tipo de usuário
    if request.user.is_admin():
        fields = Field.objects.all().order_by('name')
        churches = Church.objects.all().order_by('name')
    else:
        # Tesoureiros veem apenas seus campos e suas igrejas
        if request.user.fields.exists():
            fields = request.user.fields.all().order_by('name')
            churches = Church.objects.filter(field__in=request.user.fields.all()).order_by('name')
        else:
            fields = Field.objects.none()
            churches = Church.objects.none()
    
    context = {
        'title': 'Visualizar Transação',
        'form': form,
        'transaction': transaction,
        'readonly': True,
        'categories': Category.objects.all().order_by('name'),
        'fields': fields,
        'churches': churches,
    }
    
    return render(request, 'pages/transaction_form.html', context)

@admin_required
def transaction_edit(request, pk):
    """Editar transação - Apenas administradores"""
    transaction = get_object_or_404(Transaction, pk=pk)
    
    if request.method == 'POST':
        form = TransactionForm(request.POST, request.FILES, instance=transaction, user=request.user)
        if form.is_valid():
            form.save()
            
            # Verificar se deve criar um lembrete
            if form.cleaned_data.get('create_reminder'):
                try:
                    # Criar notificação de lembrete (1 dia após a data da transação, às 9h)
                    reminder_date = datetime.combine(transaction.date, datetime.min.time().replace(hour=9, minute=0))
                    reminder_date += timedelta(days=1)
                    reminder_date = timezone.make_aware(reminder_date)
                    
                    Notification.objects.create(
                        title=f"Lembrete: {transaction.category.name}",
                        body=f"Lembrete sobre a transação de {transaction.get_type_display().lower()} de {transaction.get_formatted_value()} para {transaction.church.name}.",
                        date=reminder_date,
                        is_read=False,
                        repeat=False,
                        repeat_frequency='none',
                        created_by=request.user
                    )
                    messages.success(request, 'Transação atualizada com sucesso e lembrete criado!')
                except Exception as e:
                    messages.success(request, 'Transação atualizada com sucesso, mas houve um erro ao criar o lembrete.')
                    print(f"Erro ao criar lembrete: {e}")
            else:
                messages.success(request, 'Transação atualizada com sucesso!')
            
            return redirect('transaction_list')
    else:
        form = TransactionForm(instance=transaction, user=request.user)
    

    
    # Preparar dados para os filtros baseado no tipo de usuário
    if request.user.is_admin():
        fields = Field.objects.all().order_by('name')
        churches = Church.objects.all().order_by('name')
    else:
        # Tesoureiros veem apenas seus campos e suas igrejas
        if request.user.fields.exists():
            fields = request.user.fields.all().order_by('name')
            churches = Church.objects.filter(field__in=request.user.fields.all()).order_by('name')
        else:
            fields = Field.objects.none()
            churches = Church.objects.none()
    
    context = {
        'title': 'Editar Transação',
        'form': form,
        'transaction': transaction,
        'categories': Category.objects.all().order_by('name'),
        'fields': fields,
        'churches': churches,
    }
    
    return render(request, 'pages/transaction_form.html', context)

@password_changed_required
@admin_required
def transaction_delete(request, pk):
    """Excluir transação - Apenas administradores"""
    transaction = get_object_or_404(Transaction, pk=pk)
    
    if request.method == 'POST':
        transaction.delete()
        messages.success(request, 'Transação excluída com sucesso!')
        return redirect('transaction_list')
    
    context = {
        'title': 'Excluir Transação',
        'transaction': transaction,
    }
    
    return render(request, 'pages/transaction_confirm_delete.html', context)

# Views de Categorias (apenas admin)
@password_changed_required
@admin_required
def category_list(request):
    """Lista de categorias"""
    categories = Category.objects.all()
    
    # Busca por nome
    search_query = request.GET.get('search', '').strip()
    if search_query:
        categories = categories.filter(name__icontains=search_query)
    
    context = {
        'title': 'Categorias',
        'categories': categories,
        'search_query': search_query,
    }
    
    return render(request, 'pages/category_list.html', context)

@password_changed_required
@admin_required
def category_create(request):
    """Criar nova categoria"""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoria criada com sucesso!')
            return redirect('category_list')
    else:
        form = CategoryForm()
    
    context = {
        'title': 'Nova Categoria',
        'form': form,
    }
    
    return render(request, 'pages/category_form.html', context)

@password_changed_required
@admin_required
def category_edit(request, pk):
    """Editar categoria"""
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoria atualizada com sucesso!')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    
    context = {
        'title': 'Editar Categoria',
        'form': form,
        'category': category,
    }
    
    return render(request, 'pages/category_form.html', context)

@password_changed_required
@admin_required
def category_delete(request, pk):
    """Excluir categoria"""
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Categoria excluída com sucesso!')
        return redirect('category_list')
    
    context = {
        'title': 'Excluir Categoria',
        'category': category,
    }
    
    return render(request, 'pages/category_confirm_delete.html', context)

# Views de Igrejas (apenas admin)
@password_changed_required
@admin_required
def church_list(request):
    """Lista de igrejas"""
    churches = Church.objects.all()
    
    # Busca por nome ou pastor
    search_query = request.GET.get('search', '').strip()
    if search_query:
        churches = churches.filter(
            Q(name__icontains=search_query) | 
            Q(shepherd__name__icontains=search_query)
        )
    
    # Filtro por campo
    field_filter = request.GET.get('field')
    selected_field_obj = None
    
    if field_filter:
        churches = churches.filter(field_id=field_filter)
        # Buscar o objeto do campo selecionado para mostrar o nome
        try:
            selected_field_obj = Field.objects.get(id=field_filter)
        except Field.DoesNotExist:
            pass
    
    # Buscar todos os campos para o filtro
    fields = Field.objects.all().order_by('name')
    
    context = {
        'title': 'Igrejas',
        'churches': churches,
        'fields': fields,
        'selected_field': field_filter,
        'selected_field_obj': selected_field_obj,
        'search_query': search_query,
    }
    
    return render(request, 'pages/church_list.html', context)

@password_changed_required
@admin_required
def church_create(request):
    """Criar nova igreja"""
    if request.method == 'POST':
        form = ChurchForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Igreja criada com sucesso!')
            return redirect('church_list')
    else:
        form = ChurchForm()
    
    context = {
        'title': 'Nova Igreja',
        'form': form,
        'shepherds': Shepherd.objects.all().order_by('name'),
        'fields': Field.objects.all().order_by('name'),
    }
    
    return render(request, 'pages/church_form.html', context)

@password_changed_required
@admin_required
def church_edit(request, pk):
    """Editar igreja"""
    church = get_object_or_404(Church, pk=pk)
    
    if request.method == 'POST':
        form = ChurchForm(request.POST, instance=church)
        if form.is_valid():
            form.save()
            messages.success(request, 'Igreja atualizada com sucesso!')
            return redirect('church_list')
    else:
        form = ChurchForm(instance=church)
    
    context = {
        'title': 'Editar Igreja',
        'form': form,
        'church': church,
        'shepherds': Shepherd.objects.all().order_by('name'),
        'fields': Field.objects.all().order_by('name'),
    }
    
    return render(request, 'pages/church_form.html', context)

@password_changed_required
@admin_required
def church_delete(request, pk):
    """Excluir igreja"""
    church = get_object_or_404(Church, pk=pk)
    
    if request.method == 'POST':
        church.delete()
        messages.success(request, 'Igreja excluída com sucesso!')
        return redirect('church_list')
    
    context = {
        'title': 'Excluir Igreja',
        'church': church,
    }
    
    return render(request, 'pages/church_confirm_delete.html', context)

# Views de Usuários (apenas admin)
@password_changed_required
@admin_required
def user_list(request):
    """Lista de usuários"""
    users = User.objects.all()
    
    # Excluir o usuário vwtechdev@gmail.com da lista
    users = users.exclude(email='vwtechdev@gmail.com')
    
    # Busca por nome, email ou função
    search_query = request.GET.get('search', '').strip()
    if search_query:
        users = users.filter(
            Q(first_name__icontains=search_query) | 
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(role__icontains=search_query)
        )
    
    context = {
        'title': 'Usuários',
        'users': users,
        'search_query': search_query,
    }
    
    return render(request, 'pages/user_list.html', context)

@password_changed_required
@admin_required
def user_create(request):
    """Criar novo usuário"""
    if request.method == 'POST':
        form = UserForm(request.POST, user=request.user)
        if form.is_valid():
            user = form.save(commit=False)
            user.save()
            
            # Salvar o formulário para processar campos many-to-many se houver
            form.save()
            
            messages.success(request, f'Usuário {user.get_full_name()} criado com sucesso!')
            return redirect('user_list')
    else:
        form = UserForm(user=request.user)
    
    context = {
        'title': 'Novo Usuário',
        'form': form,
    }
    
    return render(request, 'pages/user_form.html', context)

@password_changed_required
@admin_required
def user_edit(request, pk):
    """Editar usuário"""
    user = get_object_or_404(User, pk=pk)
    
    # Impedir edição do usuário vwtechdev@gmail.com
    if user.email == 'vwtechdev@gmail.com':
        messages.error(request, 'Este usuário não pode ser editado.')
        return redirect('user_list')
    
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuário atualizado com sucesso!')
            return redirect('user_list')
    else:
        form = UserForm(instance=user, user=request.user)
    
    context = {
        'title': 'Editar Usuário',
        'form': form,
        'user_obj': user,
    }
    
    return render(request, 'pages/user_form.html', context)

@password_changed_required
@admin_required
def user_delete(request, pk):
    """Desativar usuário"""
    user = get_object_or_404(User, pk=pk)
    
    # Impedir desativação do usuário vwtechdev@gmail.com
    if user.email == 'vwtechdev@gmail.com':
        messages.error(request, 'Este usuário não pode ser desativado.')
        return redirect('user_list')
    
    if request.method == 'POST':
        user.is_active = False
        user.save()
        messages.success(request, 'Usuário desativado com sucesso!')
        return redirect('user_list')
    
    context = {
        'title': 'Desativar Usuário',
        'user_obj': user,
    }
    
    return render(request, 'pages/user_confirm_delete.html', context)

@admin_required
def user_reset_password(request, pk):
    """Resetar senha do usuário para a senha padrão"""
    if request.method == 'POST':
        user = get_object_or_404(User, pk=pk)
        
        # Impedir reset de senha do usuário vwtechdev@gmail.com
        if user.email == 'vwtechdev@gmail.com':
            return JsonResponse({'success': False, 'error': 'Este usuário não pode ter a senha resetada.'})
        
        try:
            user.password = make_password('nations123456')
            user.password_changed = False  # Força o usuário a trocar a senha novamente
            user.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método não permitido'})



@password_changed_required
@admin_required
def user_activate(request, pk):
    """Reativar usuário"""
    user = get_object_or_404(User, pk=pk)
    
    # Impedir ativação do usuário vwtechdev@gmail.com
    if user.email == 'vwtechdev@gmail.com':
        messages.error(request, 'Este usuário não pode ser reativado.')
        return redirect('user_list')
    
    if request.method == 'POST':
        user.is_active = True
        user.save()
        messages.success(request, 'Usuário reativado com sucesso!')
        return redirect('user_list')
    
    context = {
        'title': 'Reativar Usuário',
        'user_obj': user,
    }
    
    return render(request, 'pages/user_confirm_delete.html', context)

# Views de Campos (apenas admin)
@password_changed_required
@admin_required
def field_list(request):
    """Lista de campos"""
    
    fields = Field.objects.annotate(
        church_count=Count('church', distinct=True)
    ).all()
    
    # Busca por nome
    search_query = request.GET.get('search', '').strip()
    if search_query:
        fields = fields.filter(name__icontains=search_query)
    
    context = {
        'title': 'Campos',
        'fields': fields,
        'search_query': search_query,
    }
    
    return render(request, 'pages/field_list.html', context)

@password_changed_required
@admin_required
def field_create(request):
    """Criar novo campo"""
    if request.method == 'POST':
        form = FieldForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Campo criado com sucesso!')
            return redirect('field_list')
    else:
        form = FieldForm()
    
    context = {
        'title': 'Novo Campo',
        'form': form,
    }
    
    return render(request, 'pages/field_form.html', context)

@password_changed_required
@admin_required
def field_edit(request, pk):
    """Editar campo"""
    field = get_object_or_404(Field, pk=pk)
    
    if request.method == 'POST':
        form = FieldForm(request.POST, instance=field)
        if form.is_valid():
            form.save()
            messages.success(request, 'Campo atualizado com sucesso!')
            return redirect('field_list')
    else:
        form = FieldForm(instance=field)
    
    context = {
        'title': 'Editar Campo',
        'form': form,
        'field': field,
    }
    
    return render(request, 'pages/field_form.html', context)

@password_changed_required
@admin_required
def field_delete(request, pk):
    """Excluir campo"""
    field = get_object_or_404(Field, pk=pk)
    
    if request.method == 'POST':
        field.delete()
        messages.success(request, 'Campo excluído com sucesso!')
        return redirect('field_list')
    
    context = {
        'title': 'Excluir Campo',
        'field': field,
    }
    
    return render(request, 'pages/field_confirm_delete.html', context)

# API Views para AJAX
@password_changed_required
def get_churches(request):
    """Retorna igrejas de um campo e/ou pastor via AJAX"""
    field_id = request.GET.get('field')
    shepherd_id = request.GET.get('shepherd')
    
    # Base segura conforme permissões
    if request.user.is_admin():
        churches = Church.objects.all()
    else:
        if request.user.fields.exists():
            churches = Church.objects.filter(field__in=request.user.fields.all())
        else:
            churches = Church.objects.none()
    
    if field_id:
        churches = churches.filter(field_id=field_id)
    
    if shepherd_id:
        churches = churches.filter(shepherd_id=shepherd_id)
    
    churches_data = churches.values('id', 'name')
    return JsonResponse({'churches': list(churches_data)})



@password_changed_required
@admin_or_treasurer_required
def transaction_export_pdf(request):
    """Exporta transações filtradas para PDF"""
    
    # Debug: Log dos parâmetros recebidos
    print(f"DEBUG - Parâmetros recebidos na exportação PDF:")
    print(f"DEBUG - category: {request.GET.get('category', '')}")
    print(f"DEBUG - type: {request.GET.get('type', '')}")
    print(f"DEBUG - date_from: {request.GET.get('date_from', '')}")
    print(f"DEBUG - date_to: {request.GET.get('date_to', '')}")
    print(f"DEBUG - field: {request.GET.get('field', '')}")
    print(f"DEBUG - church: {request.GET.get('church', '')}")
    print(f"DEBUG - search: {request.GET.get('search', '')}")
    
    # Obter os mesmos filtros da view transaction_list
    if request.user.is_admin():
        transactions = Transaction.objects.all()
    elif request.user.is_supervisor():
        transactions = get_transactions_for_user(request.user)
    else:
        # Tesoureiro: exportar apenas transações criadas por ele
        transactions = Transaction.objects.filter(user=request.user)
    
    # Aplicar filtros
    search = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    transaction_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    selected_field = request.GET.get('field', '')
    selected_church = request.GET.get('church', '')
    selected_shepherd = request.GET.get('shepherd', '')
    selected_user = request.GET.get('user', '')
    
    # Debug: Log dos filtros aplicados
    print(f"DEBUG - Filtros aplicados:")
    print(f"DEBUG - Total inicial de transações: {transactions.count()}")
    
    # Definir datas padrão se não fornecidas
    today = date.today()
    if not date_from:
        first_day = date(today.year, today.month, 1)
        date_from = first_day.strftime('%Y-%m-%d')
    
    if not date_to:
        last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])
        date_to = last_day.strftime('%Y-%m-%d')
    
    # Aplicar filtros
    if search:
        transactions = transactions.filter(
            Q(desc__icontains=search) | 
            Q(category__name__icontains=search)
        )
        print(f"DEBUG - Após filtro de busca: {transactions.count()} transações")
    
    if category_id:
        transactions = transactions.filter(category_id=category_id)
        print(f"DEBUG - Após filtro de categoria: {transactions.count()} transações")
    
    if transaction_type:
        transactions = transactions.filter(type=transaction_type)
        print(f"DEBUG - Após filtro de tipo: {transactions.count()} transações")
    
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
        print(f"DEBUG - Após filtro de data inicial: {transactions.count()} transações")
    
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
        print(f"DEBUG - Após filtro de data final: {transactions.count()} transações")
    
    if selected_field:
        if request.user.is_admin():
            transactions = transactions.filter(church__field_id=selected_field)
        elif request.user.fields.count() > 1:
            if request.user.fields.filter(id=selected_field).exists():
                transactions = transactions.filter(church__field_id=selected_field)
        print(f"DEBUG - Após filtro de campo: {transactions.count()} transações")
    
    if selected_church:
        transactions = transactions.filter(church_id=selected_church)
        print(f"DEBUG - Após filtro de igreja: {transactions.count()} transações")
    
    if selected_shepherd:
        transactions = transactions.filter(church__shepherd_id=selected_shepherd)
        print(f"DEBUG - Após filtro de pastor: {transactions.count()} transações")
    
    if selected_user and request.user.is_admin():
        transactions = transactions.filter(user_id=selected_user)
        print(f"DEBUG - Após filtro de usuário: {transactions.count()} transações")
    
    print(f"DEBUG - Total final de transações: {transactions.count()}")
    
    # Calcular totais
    total_transactions = transactions.count()
    total_income = transactions.filter(type='income').aggregate(total=Sum('value'))['total'] or 0
    total_expense = transactions.filter(type='expense').aggregate(total=Sum('value'))['total'] or 0
    balance = total_income - total_expense
    
    # Criar o PDF
    response = HttpResponse(content_type='application/pdf')
    # Formatar nome do arquivo no formato brasileiro (dia-mes-ano)
    filename_date = date.today().strftime("%d-%m-%Y")
    response['Content-Disposition'] = f'attachment; filename="transacoes_{filename_date}.pdf"'
    
    # Configurar o documento A4 em retrato com margens de 0.5 polegadas
    doc = SimpleDocTemplate(response, pagesize=A4, 
                           leftMargin=0.5*inch, rightMargin=0.5*inch, 
                           topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#673ab7')
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=10,
        spaceAfter=8,
        alignment=TA_LEFT,
        textColor=colors.HexColor('#495057')
    )
    
    # Título com logo
    title_data = []
    
    # Verificar se o logo existe
    possible_paths = [
        os.path.join(settings.STATIC_ROOT, 'img', 'icon.png'),
        os.path.join(settings.BASE_DIR, 'static', 'img', 'icon.png'),
        os.path.join(os.path.dirname(__file__), 'static', 'img', 'icon.png'),
        os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'img', 'icon.png'),
    ]
    
    icon_path = None
    icon_exists = False
    
    for path in possible_paths:
        if os.path.exists(path):
            icon_path = path
            icon_exists = True
            break
    
    if icon_exists:
        try:
            # Carregar o logo PNG
            icon = Image(icon_path, width=1.2*inch, height=1.2*inch)
            title_data = [[icon, Paragraph("Relatório de Transações", title_style)]]
        except Exception as e:
            print(f"Erro ao carregar logo PNG: {e}")
            icon_symbol = Paragraph("●", ParagraphStyle(
                'IconStyle',
                parent=styles['Normal'],
                fontSize=40,
                alignment=TA_CENTER,
                textColor=colors.HexColor('#673ab7')
            ))
            title_data = [[icon_symbol, Paragraph("Relatório de Transações", title_style)]]
    else:
        icon_symbol = Paragraph("●", ParagraphStyle(
            'IconStyle',
            parent=styles['Normal'],
            fontSize=40,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#673ab7')
        ))
        title_data = [[icon_symbol, Paragraph("Relatório de Transações", title_style)]]
    
    # Criar tabela do título
    title_table = Table(title_data, colWidths=[1.5*inch, 5.5*inch])
    title_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    
    elements.append(title_table)
    elements.append(Spacer(1, 15))
    
    # Mostrar filtros aplicados
    filters_applied = []
    
    # Formatar datas para o formato brasileiro
    try:
        date_from_formatted = datetime.strptime(date_from, '%Y-%m-%d').strftime('%d/%m/%Y')
    except (ValueError, TypeError):
        date_from_formatted = date_from
    
    try:
        date_to_formatted = datetime.strptime(date_to, '%Y-%m-%d').strftime('%d/%m/%Y')
    except (ValueError, TypeError):
        date_to_formatted = date_to
    
    filters_applied.append(f"Período: {date_from_formatted} a {date_to_formatted}")
    
    if category_id:
        try:
            category_name = Category.objects.get(id=category_id).name
            filters_applied.append(f"Categoria: {category_name}")
        except Category.DoesNotExist:
            pass
    
    if transaction_type:
        type_name = "Entrada" if transaction_type == "income" else "Saída"
        filters_applied.append(f"Tipo: {type_name}")
    
    if selected_field:
        try:
            field_name = Field.objects.get(id=selected_field).name
            filters_applied.append(f"Campo: {field_name}")
        except Field.DoesNotExist:
            pass
    
    if selected_church:
        try:
            church_name = Church.objects.get(id=selected_church).name
            filters_applied.append(f"Igreja: {church_name}")
        except Church.DoesNotExist:
            pass
    
    if selected_shepherd:
        try:
            shepherd_name = Shepherd.objects.get(id=selected_shepherd).name
            filters_applied.append(f"Pastor: {shepherd_name}")
        except Shepherd.DoesNotExist:
            pass
    
    if search:
        filters_applied.append(f"Busca: {search}")
    
    # Adicionar filtros aplicados
    if filters_applied:
        elements.append(Paragraph("Filtros Aplicados:", subtitle_style))
        for filter_info in filters_applied:
            elements.append(Paragraph(f"• {filter_info}", subtitle_style))
        elements.append(Spacer(1, 10))
    
    # Informações do relatório
    report_info = [
        f"Total de Transações: {total_transactions}",
        f"Total de Entradas: R$ {total_income:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
        f"Total de Saídas: R$ {total_expense:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
        f"Saldo: R$ {balance:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
    ]
    
    for info in report_info:
        elements.append(Paragraph(info, subtitle_style))
    
    elements.append(Spacer(1, 15))
    
    # Cabeçalho da tabela com todas as colunas solicitadas
    headers = ['Data', 'Tipo', 'Categoria', 'Campo', 'Igreja', 'Descrição', 'Valor', 'Usuário', 'Pastor']
    data = [headers]
    
    # Dados das transações
    for transaction in transactions.order_by('-date').select_related('category', 'church', 'church__field', 'church__shepherd', 'user'):
        # Formatar valor monetário
        formatted_value = f"{transaction.value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        
        # Estilos para células
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=7,
            leading=9,
            alignment=TA_LEFT,
            leftIndent=1,
            rightIndent=1,
            spaceBefore=1,
            spaceAfter=1
        )
        
        value_style = ParagraphStyle(
            'ValueStyle',
            parent=styles['Normal'],
            fontSize=7,
            leading=9,
            alignment=TA_RIGHT,
            leftIndent=1,
            rightIndent=1,
            spaceBefore=1,
            spaceAfter=1
        )
        
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=7,
            leading=9,
            alignment=TA_CENTER,
            leftIndent=1,
            rightIndent=1,
            spaceBefore=1,
            spaceAfter=1
        )
        
        type_style = ParagraphStyle(
            'TypeStyle',
            parent=styles['Normal'],
            fontSize=7,
            leading=9,
            alignment=TA_CENTER,
            leftIndent=1,
            rightIndent=1,
            spaceBefore=1,
            spaceAfter=1
        )
        
        row = [
            Paragraph(transaction.date.strftime('%d/%m/%Y'), date_style),
            Paragraph('Entrada' if transaction.type == 'income' else 'Saída', type_style),
            Paragraph(transaction.category.name, cell_style),
            Paragraph(transaction.church.field.name, cell_style),
            Paragraph(transaction.church.name, cell_style),
            Paragraph(transaction.desc or '-', cell_style),
            Paragraph(formatted_value, value_style),
            Paragraph(transaction.user.get_full_name() if transaction.user else '-', cell_style),
            Paragraph(transaction.church.shepherd.name if transaction.church.shepherd else '-', cell_style)
        ]
        data.append(row)
    
    # Criar tabela com larguras otimizadas para retrato (9 colunas)
    table = Table(data, colWidths=[0.7*inch, 0.5*inch, 0.8*inch, 0.8*inch, 1.0*inch, 1.2*inch, 0.7*inch, 0.8*inch, 0.8*inch])
    
    # Estilo da tabela
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#673ab7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ])
    
    # Alternar cores das linhas
    for i in range(1, len(data)):
        if i % 2 == 0:
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Gerar PDF
    doc.build(elements)
    
    return response


@password_changed_required
@admin_or_treasurer_required
def transaction_export_xlsx(request):
    """Exporta transações filtradas para XLSX"""
    
    # Debug: Log dos parâmetros recebidos
    print(f"DEBUG - Parâmetros recebidos na exportação XLSX:")
    print(f"DEBUG - category: {request.GET.get('category', '')}")
    print(f"DEBUG - type: {request.GET.get('type', '')}")
    print(f"DEBUG - date_from: {request.GET.get('date_from', '')}")
    print(f"DEBUG - date_to: {request.GET.get('date_to', '')}")
    print(f"DEBUG - field: {request.GET.get('field', '')}")
    print(f"DEBUG - church: {request.GET.get('church', '')}")
    print(f"DEBUG - shepherd: {request.GET.get('shepherd', '')}")
    
    # Obter os mesmos filtros da view transaction_list
    if request.user.is_admin():
        base_transactions = Transaction.objects.all()
    elif request.user.is_supervisor():
        base_transactions = get_transactions_for_user(request.user)
    else:
        # Tesoureiro: exportar apenas transações criadas por ele
        base_transactions = Transaction.objects.filter(user=request.user)
    
    # Aplicar filtros
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    selected_category = request.GET.get('category', '')
    selected_type = request.GET.get('type', '')
    selected_field = request.GET.get('field', '')
    selected_church = request.GET.get('church', '')
    selected_shepherd = request.GET.get('shepherd', '')
    selected_user = request.GET.get('user', '')
    
    # Definir datas padrão se não fornecidas
    today = date.today()
    if not date_from:
        first_day = date(today.year, today.month, 1)
        date_from = first_day.strftime('%Y-%m-%d')
    
    if not date_to:
        last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])
        date_to = last_day.strftime('%Y-%m-%d')
    
    # Debug: Log dos filtros aplicados
    print(f"DEBUG - Total inicial de transações: {base_transactions.count()}")
    
    # Aplicar filtros
    filtered_transactions = base_transactions
    if search:
        filtered_transactions = filtered_transactions.filter(
            Q(desc__icontains=search) |
            Q(category__name__icontains=search)
        )
        print(f"DEBUG - Após filtro de busca: {filtered_transactions.count()} transações")

    
    if date_from:
        filtered_transactions = filtered_transactions.filter(date__gte=date_from)
        print(f"DEBUG - Após filtro de data inicial: {filtered_transactions.count()} transações")
    
    if date_to:
        filtered_transactions = filtered_transactions.filter(date__lte=date_to)
        print(f"DEBUG - Após filtro de data final: {filtered_transactions.count()} transações")
    
    if selected_category:
        filtered_transactions = filtered_transactions.filter(category_id=selected_category)
        print(f"DEBUG - Após filtro de categoria: {filtered_transactions.count()} transações")
    
    if selected_type:
        filtered_transactions = filtered_transactions.filter(type=selected_type)
        print(f"DEBUG - Após filtro de tipo: {filtered_transactions.count()} transações")
    
    if selected_field:
        filtered_transactions = filtered_transactions.filter(church__field_id=selected_field)
        print(f"DEBUG - Após filtro de campo: {filtered_transactions.count()} transações")
    
    if selected_church:
        filtered_transactions = filtered_transactions.filter(church_id=selected_church)
        print(f"DEBUG - Após filtro de igreja: {filtered_transactions.count()} transações")
    
    if selected_shepherd:
        filtered_transactions = filtered_transactions.filter(church__shepherd_id=selected_shepherd)
        print(f"DEBUG - Após filtro de pastor: {filtered_transactions.count()} transações")
    
    if selected_user and request.user.is_admin():
        filtered_transactions = filtered_transactions.filter(user_id=selected_user)
        print(f"DEBUG - Após filtro de usuário: {filtered_transactions.count()} transações")
    
    print(f"DEBUG - Total final de transações: {filtered_transactions.count()}")
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transações"
    
    # Cabeçalho
    colunas = [
        "ID", "Data", "Categoria", "Tipo", "Valor", "Igreja", 
        "Campo", "Pastor", "Descrição", "Usuário", "Data de Criação"
    ]
    ws.append(colunas)
    
    # Estilizar cabeçalho
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="673AB7", end_color="673AB7", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Dados do ORM
    for transaction in filtered_transactions.select_related('category', 'church', 'church__field', 'church__shepherd', 'user'):
        ws.append([
            transaction.id,
            transaction.date.strftime('%d/%m/%Y'),
            transaction.category.name if transaction.category else '',
            'Entrada' if transaction.type == 'income' else 'Saída',
            f'R$ {transaction.value:.2f}',
            transaction.church.name if transaction.church else '',
            transaction.church.field.name if transaction.church and transaction.church.field else '',
            transaction.church.shepherd.name if transaction.church and transaction.church.shepherd else '',
            transaction.desc or '',
            (transaction.user.get_full_name() or transaction.user.username) if transaction.user else '',
            transaction.created_at.strftime('%d/%m/%Y %H:%M') if transaction.created_at else ''
        ])
    
    # Ajustar largura das colunas
    column_widths = [8, 12, 20, 10, 15, 25, 20, 20, 30, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Resposta HTTP para download
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename_date = date.today().strftime("%d-%m-%Y")
    response["Content-Disposition"] = f'attachment; filename="transacoes_{filename_date}.xlsx"'
    wb.save(response)
    
    return response


@password_changed_required
def churches_by_field_api(request, field_id):
    """API para retornar igrejas de um campo específico"""
    
    try:
        # Verificar permissão do campo para tesoureiros
        if request.user.is_admin():
            field = Field.objects.get(pk=field_id)
        else:
            field = Field.objects.get(pk=field_id)
            if not request.user.fields.filter(pk=field.pk).exists():
                return JsonResponse({'error': 'Sem permissão para este campo'}, status=403)
        
        # Buscar igrejas do campo
        if request.user.is_admin():
            churches = Church.objects.filter(field=field).order_by('name')
        else:
            churches = Church.objects.filter(field=field).order_by('name')
        
        churches_data = []
        for church in churches:
            churches_data.append({
                'id': church.id,
                'name': church.name,
                'address': church.address or '',
                'shepherd': church.shepherd.name if church.shepherd else ''
            })
        
        return JsonResponse({
            'field': field.name,
            'churches': churches_data
        })
        
    except Field.DoesNotExist:
        return JsonResponse({'error': 'Campo não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@password_changed_required
def shepherds_by_field_api(request, field_id):
    """API para retornar pastores de um campo específico"""
    
    try:
        # Verificar permissão do campo para tesoureiros
        if request.user.is_admin():
            field = Field.objects.get(pk=field_id)
        else:
            field = Field.objects.get(pk=field_id)
            if not request.user.fields.filter(pk=field.pk).exists():
                return JsonResponse({'error': 'Sem permissão para este campo'}, status=403)
        
        # Buscar pastores do campo (distinct para evitar duplicatas)
        shepherds = Shepherd.objects.filter(church__field=field).distinct().order_by('name')
        
        shepherds_data = []
        for shepherd in shepherds:
            shepherds_data.append({
                'id': shepherd.id,
                'name': shepherd.name
            })
        
        return JsonResponse({
            'field': field.name,
            'shepherds': shepherds_data
        })
        
    except Field.DoesNotExist:
        return JsonResponse({'error': 'Campo não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@password_changed_required
def category_info_api(request, category_id):
    """API para retornar informações de uma categoria específica"""
    try:
        category = get_object_or_404(Category, pk=category_id)
        
        return JsonResponse({
            'id': category.id,
            'name': category.name,
            'mandatory_proof': category.mandatory_proof
        })
        
    except Category.DoesNotExist:
        return JsonResponse({'error': 'Categoria não encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Views de Pastores (apenas admin)
@password_changed_required
@admin_required
def shepherd_list(request):
    """Lista de pastores"""
    shepherds = Shepherd.objects.all().order_by('name')
    
    # Busca por nome
    search_query = request.GET.get('search', '').strip()
    if search_query:
        shepherds = shepherds.filter(name__icontains=search_query)
    
    # Incluir igrejas vinculadas para cada pastor
    shepherds = shepherds.prefetch_related('church_set')
    
    context = {
        'title': 'Pastores',
        'shepherds': shepherds,
        'search_query': search_query,
    }
    
    return render(request, 'pages/shepherd_list.html', context)

@password_changed_required
@admin_required
def shepherd_create(request):
    """Criar novo pastor"""
    if request.method == 'POST':
        form = ShepherdForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pastor criado com sucesso!')
            return redirect('shepherd_list')
    else:
        form = ShepherdForm()
    
    context = {
        'title': 'Novo Pastor',
        'form': form,
    }
    
    return render(request, 'pages/shepherd_form.html', context)

@password_changed_required
@admin_required
def shepherd_edit(request, pk):
    """Editar pastor"""
    shepherd = get_object_or_404(Shepherd, pk=pk)
    
    if request.method == 'POST':
        form = ShepherdForm(request.POST, instance=shepherd)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pastor atualizado com sucesso!')
            return redirect('shepherd_list')
    else:
        form = ShepherdForm(instance=shepherd)
    
    # Calcular estatísticas
    churches = shepherd.church_set.all()
    unique_fields = churches.values('field').distinct().count()
    
    context = {
        'title': 'Editar Pastor',
        'form': form,
        'shepherd': shepherd,
        'church_count': churches.count(),
        'unique_fields_count': unique_fields,
    }
    
    return render(request, 'pages/shepherd_form.html', context)

@password_changed_required
@admin_required
def shepherd_delete(request, pk):
    """Excluir pastor"""
    shepherd = get_object_or_404(Shepherd, pk=pk)
    
    if request.method == 'POST':
        shepherd.delete()
        messages.success(request, f'Pastor {shepherd.name} excluído com sucesso!')
        return redirect('shepherd_list')
    
    context = {
        'title': 'Excluir Pastor',
        'shepherd': shepherd,
    }
    
    return render(request, 'pages/shepherd_confirm_delete.html', context)

# Views de Logs de Acesso (apenas admin)
@password_changed_required
@admin_required
def access_log_list(request):
    """Lista de logs de acesso - apenas do primeiro ao último dia do mês atual"""
    from datetime import date
    import calendar
    
    # Calcular primeiro e último dia do mês atual
    today = date.today()
    first_day_of_month = today.replace(day=1)
    last_day_of_month = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    
    # Filtrar logs apenas do mês atual
    logs = AccessLog.objects.select_related('user').exclude(
        user__email='vwtechdev@gmail.com'
    ).filter(
        timestamp__date__gte=first_day_of_month,
        timestamp__date__lte=last_day_of_month
    ).order_by('-timestamp')
    
    # Busca por nome ou email do usuário
    search_query = request.GET.get('search', '').strip()
    if search_query:
        logs = logs.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    # Filtro por data (opcional, mas limitado ao mês atual)
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            # Garantir que não vá antes do primeiro dia do mês
            if date_from_obj < first_day_of_month:
                date_from_obj = first_day_of_month
            logs = logs.filter(timestamp__date__gte=date_from_obj)
        except ValueError:
            pass  # Ignora datas inválidas
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            # Garantir que não vá depois do último dia do mês
            if date_to_obj > last_day_of_month:
                date_to_obj = last_day_of_month
            logs = logs.filter(timestamp__date__lte=date_to_obj)
        except ValueError:
            pass  # Ignora datas inválidas
    
    context = {
        'title': 'Logs de Acesso',
        'logs': logs,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'first_day_of_month': first_day_of_month,
        'last_day_of_month': last_day_of_month,
        'current_month': today.strftime('%B de %Y'),
    }
    
    return render(request, 'pages/access_log_list.html', context)


# Views de Notificações (apenas admin)
@password_changed_required
@admin_required
def notification_list(request):
    """Lista de notificações - apenas as criadas pelo usuário logado"""
    notifications = Notification.objects.filter(
        created_by=request.user
    ).select_related('created_by').order_by('-created_at')
    
    # Busca por título ou mensagem
    search_query = request.GET.get('search', '').strip()
    if search_query:
        notifications = notifications.filter(
            Q(title__icontains=search_query) |
            Q(body__icontains=search_query)
        )
    
    context = {
        'title': 'Minhas Notificações',
        'notifications': notifications,
        'search_query': search_query,
    }
    
    return render(request, 'pages/notification_list.html', context)


@password_changed_required
@admin_required
def notification_create(request):
    """Criar nova notificação"""
    if request.method == 'POST':
        form = NotificationForm(request.POST)
        if form.is_valid():
            notification = form.save(commit=False)
            notification.created_by = request.user
            notification.save()
            
            messages.success(request, 'Notificação criada com sucesso!')
            return redirect('notification_list')
    else:
        form = NotificationForm()
    
    context = {
        'title': 'Nova Notificação',
        'form': form,
    }
    
    return render(request, 'pages/notification_form.html', context)


@password_changed_required
@admin_required
def notification_edit(request, pk):
    """Editar notificação - apenas as criadas pelo usuário logado"""
    notification = get_object_or_404(Notification, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        form = NotificationForm(request.POST, instance=notification)
        if form.is_valid():
            updated_notification = form.save(commit=False)
            updated_notification.save()
            messages.success(request, 'Notificação atualizada com sucesso!')
            return redirect('notification_list')
    else:
        form = NotificationForm(instance=notification)
    
    context = {
        'title': 'Editar Notificação',
        'form': form,
        'notification': notification,
    }
    
    return render(request, 'pages/notification_form.html', context)

@password_changed_required
@admin_required
def notification_delete(request, pk):
    """Excluir notificação - apenas as criadas pelo usuário logado"""
    notification = get_object_or_404(Notification, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        notification.delete()
        messages.success(request, 'Notificação excluída com sucesso!')
        return redirect('notification_list')
    
    context = {
        'title': 'Excluir Notificação',
        'notification': notification,
    }
    
    return render(request, 'pages/notification_confirm_delete.html', context)


@password_changed_required
@admin_required
def notification_mark_read(request, pk):
    """Marcar notificação como lida/não lida via AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            is_read = data.get('is_read', False)
            
            notification = get_object_or_404(Notification, pk=pk, created_by=request.user)
            notification.is_read = is_read
            notification.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Status atualizado com sucesso',
                'is_read': is_read
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({
        'success': False,
        'error': 'Método não permitido'
    }, status=405)

@password_changed_required
def get_today_notifications(request):
    """Buscar notificações vencidas ou do dia atual que não foram lidas via AJAX - apenas do usuário logado"""
    if request.method == 'GET':
        try:
            today = date.today()
            
            # Buscar notificações vencidas ou com data igual a atual e que não foram lidas
            # Compara apenas a data, ignorando a hora
            notifications = Notification.objects.filter(
                date__date__lte=today,  # Data menor ou igual ao dia atual (vencidas ou hoje)
                is_read=False,  # Apenas não lidas
                created_by=request.user
            ).select_related('created_by').order_by('-date')[:10]
            
            notifications_data = []
            for notification in notifications:
                notifications_data.append({
                    'id': notification.id,
                    'title': notification.title,
                    'body': notification.body,
                    'date': notification.date.strftime('%d/%m/%Y %H:%M'),
                    'created_by': notification.created_by.get_full_name() or notification.created_by.username,
                    'repeat': notification.repeat,
                    'repeat_frequency': notification.get_repeat_frequency_display() if notification.repeat else None,
                })
            
            return JsonResponse({
                'success': True,
                'notifications': notifications_data,
                'count': len(notifications_data)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'error': 'Método não permitido'
    }, status=405)

# Health Check para monitoramento
@require_http_methods(["GET"])
def health_check(request):
    """Endpoint de health check para monitoramento"""
    try:
        # Verificar conexão com o banco de dados
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
        
        return JsonResponse({
            'status': 'healthy',
            'timestamp': timezone.now().isoformat(),
            'database': 'connected'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'unhealthy',
            'timestamp': timezone.now().isoformat(),
            'error': str(e)
        }, status=500)


